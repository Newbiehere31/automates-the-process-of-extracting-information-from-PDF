import glob
from flask import Flask, request, render_template, send_from_directory
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
import PyPDF2
import openpyxl
import re
import datetime
import nltk
from openpyxl.styles import PatternFill

nltk.download('punkt')

app = Flask(__name__)

# Replace with your own credentials JSON file path
CREDENTIALS_FILE = 'test.json'

# Define the column indices for PDF links, report numbers, and the new "Status" column
PDF_LINK_COLUMN = 4  # Column D
REPORT_NUMBER_COLUMN = 3  # Column C
STATUS_COLUMN = 12  # Column L
DATE_STATUS_COLUMN = 13
DATE_COLUMN = 5  # Column E
import os
app.template_folder = os.path.join(os.path.dirname(__file__))
@app.route('/', methods=['GET', 'POST'])
def index():
    message = None  # Initialize the message as None
    if request.method == 'POST':
        excel_file = request.files['excel_file']
        if excel_file:
            excel_file.save(excel_file.filename)  # Save the uploaded Excel file
            message = process_excel_file(excel_file.filename)
            return render_template('index.html', message=message)
    return render_template('index.html', message=message)

def download_file(credentials_file, pdf_url):
    try:
        # Load the credentials from the JSON file
        creds = service_account.Credentials.from_service_account_file(credentials_file, scopes=[
            'https://www.googleapis.com/auth/drive.readonly'])

        # Create a Google Drive service instance
        service = build('drive', 'v3', credentials=creds)

        # Extract the file ID from the PDF URL
        file_id = pdf_url.split('/')[-2]

        # Request metadata for the file
        file_metadata = service.files().get(fileId=file_id).execute()

        # Create a local file path based on the file's name
        output_file = file_metadata['name']

        # Create a writable stream to save the file
        with open(output_file, 'wb') as f:
            request = service.files().get_media(fileId=file_id)
            media = MediaIoBaseDownload(f, request)

            # Download the file in chunks
            done = False
            while not done:
                _, done = media.next_chunk()

        print(f'File downloaded and saved as "{output_file}"')

        return output_file

    except Exception as e:
        print(f'Error: {e}')
        return None


def extract_text_from_pdf(pdf_file_path, max_pages=3):
    try:
        # Open the PDF file
        with open(pdf_file_path, 'rb') as pdf_file:
            # Create a PDF reader object using PdfReader
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            # Initialize an empty string to store the extracted text
            extracted_text = ''

            # Loop through each page and extract text, up to max_pages
            for page_num, page in enumerate(pdf_reader.pages):
                if page_num >= max_pages:
                    break
                extracted_text += page.extract_text()

            return extracted_text

    except Exception as e:
        print(f'Error extracting text from PDF: {e}')
        return None


def extract_dates_from_text(extracted_text):
    # Define regular expression patterns for different date formats
    date_patterns = [
        r'\b(\d{1,2}/\d{1,2}/\d{4})\b',  # dd/mm/yyyy
        r'\b(\d{4}/\d{1,2}/\d{1,2})\b',  # yyyy/mm/dd
        r'\b(\d{1,2} [A-Za-z]+ \d{4})\b',  # 9 September 2022
        r'\b(\d{1,2}-\d{1,2}-\d{4})\b',  # dd-mm-yyyy
        r'\b(0[1-9]/0[1-9]/\d{4})\b',  # 01/01/yyyy to 09/09/yyyy
        r'\b(0[1-9]/\d{1,2}/\d{4})\b',  # 01/10/yyyy to 09/12/yyyy
        r'\b(\d{1,2}/0[1-9]/\d{4})\b',  # 10/01/yyyy to 31/09/yyyy
        r'\b(\d{1,2}/\d{1,2}/\d{4})\b',  # 10/10/yyyy to 31/12/yyyy
    ]

    extracted_dates = []

    for pattern in date_patterns:
        dates = re.findall(pattern, extracted_text.replace(' ',''))
        extracted_dates.extend(dates)
    print(extracted_dates)
    return extracted_dates


def process_excel_file(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the first worksheet
    worksheet = workbook.active

    # Add a new "Status" column with a header in cell L1
    worksheet.cell(row=1, column=STATUS_COLUMN, value="Status")
    worksheet.cell(row=1, column=DATE_STATUS_COLUMN, value="Date Status")

    # Iterate through rows starting from row 3 and column D (PDF links)
    for row_index, row in enumerate(
            worksheet.iter_rows(min_row=3, min_col=PDF_LINK_COLUMN, max_col=PDF_LINK_COLUMN, values_only=True),
            start=3):
        pdf_url_cell_value = row[0]

        # Extract the report number from column C (Report Number)
        report_number_cell_value = worksheet.cell(row=row_index, column=REPORT_NUMBER_COLUMN).value

        # Initialize status as "Not Matched"
        status = "Not Matched"
        dstatus = "Not Matched"
        # Extract data from the Google Drive link
        downloaded_file_path = download_file(CREDENTIALS_FILE, pdf_url_cell_value)
        if downloaded_file_path:
            extracted_text = extract_text_from_pdf(downloaded_file_path)
            if extracted_text:
                # Remove spaces from the extracted text
                extracted_text = extracted_text.replace(' ', '')

            report_number_cell_value = report_number_cell_value.replace(' ', '')
            print(report_number_cell_value)
            # Use regular expression to find report numbers in the extracted text

            report_number_pattern = r'(\b[\w.-]+\b|[A-Z\d_ ]+\.[\d]+)'
            # Adjust the pattern based on your report number format
            report_number_cell_value = re.findall(report_number_pattern, report_number_cell_value.strip())
            report_numbers = re.findall(report_number_pattern, extracted_text.strip())

            matching_elements = []
            print(report_numbers)
            prefixes = {"ReportIdentification", "number", "GAMINGASSOCIATES"}

            for prefix in prefixes:
                report_numbers = [text.replace(prefix, '') for text in report_numbers]

            for element1 in report_numbers:
                if element1 in report_number_cell_value:
                    matching_elements.append(element1)
                    print(matching_elements)
                    break
                elif report_number_cell_value in report_numbers or report_numbers in report_number_cell_value:
                    matching_elements.append(element1)
                    print(matching_elements)
                    break

                    # for item in report_numbers:
            if len(matching_elements) >= 1:
                status = 'Matched'

            extracted_dates = extract_dates_from_text(extracted_text)
            if extracted_dates:
                # Check if any of the extracted dates match the expected date format
                for date_str in extracted_dates:
                    try:
                        # Attempt to parse the date using different date formats
                        possible_date_formats = [
                            '%d/%m/%Y',
                            '%Y/%m/%d',
                            '%d %B %Y',
                            '%d-%B-%Y',
                            '%y-%B-%d',

                        ]
                        for date_format in possible_date_formats:
                            date_obj = datetime.datetime.strptime(date_str, date_format)
                            date_status = date_obj.strftime('%d/%m/%Y')
                            dstatus = "Matched"
                            break
                    except ValueError:
                        pass  # Continue i
        worksheet.cell(row=row_index, column=STATUS_COLUMN, value=status)
        worksheet.cell(row=row_index, column=DATE_STATUS_COLUMN, value=dstatus)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        if status == "Not Matched" or dstatus == "Not Matched":
            for col in (STATUS_COLUMN, DATE_STATUS_COLUMN):
                cell = worksheet.cell(row=row_index, column=col)
                cell.fill = yellow_fill

    workbook.save(file_path)
    print(f'Status updated in "{file_path}"')


    current_directory = os.getcwd()
    pdf_pattern = os.path.join(current_directory, '*.pdf')
    pdf_files = glob.glob(pdf_pattern)

    for pdf_file in pdf_files:
        try:
            os.remove(pdf_file)
            print(f"Deleted: {pdf_file}")
        except Exception as e:
            print(f"Error deleting {pdf_file}: {e}")

    message = f'Status updated in "{file_path}"'
    return message

if __name__ == '__main__':
    app.run(debug=True)

